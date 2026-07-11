"""
tier.py
-------
Reads raw research text directly from the Excel file and uses a local LLM
(Ollama) to classify it into tier codes, following the EXACT rubric rules
already locked earlier in this project. This does not invent new judgment -
it applies rules that were already decided by hand.

NEW: skips any cell that is highlighted RED in Excel, and skips any brand
where Product_Fits_Returns_Category = N (e.g. food/beverage brands that
don't need returns tooling).

Works directly on the .xlsx file (not CSV) because color highlighting only
exists in Excel's format - CSV has no concept of cell color.

SETUP:
    ollama serve                          (in one terminal, leave running)
    $env:USE_LLM_EXPLANATIONS="true"
    python tier.py

SAFE BY DEFAULT:
    - Only fills BLANK tier cells (won't overwrite tiers you set by hand)
    - Skips any cell highlighted red
    - Skips any brand where Product_Fits_Returns_Category = N
    - If a raw text cell says "Not yet checked" / "Not yet researched" /
      is empty, there's nothing to classify - left blank, never guessed
    - If the LLM returns something that isn't a valid code, the cell is
      left blank and a warning is printed - never writes a guess

NOTE: Product_Fits_Returns_Category is NOT classified by this script - it's
set manually, since it needs real-world brand knowledge (e.g. "is this a
food brand") rather than reading research text.
"""

import os
import sys
import json
import time
import urllib.request
import urllib.error
import openpyxl

OLLAMA_URL = "http://localhost:11434/api/chat"
OLLAMA_MODEL = os.environ.get("OLLAMA_MODEL", "llama3.2")
USE_LLM_EXPLANATIONS = os.environ.get("USE_LLM_EXPLANATIONS", "").lower() == "true"
USE_LLM = USE_LLM_EXPLANATIONS

INPUT_XLSX = "CP-scorebook.xlsx"
OUTPUT_XLSX = "CP-scorebook.xlsx"  # overwrites in place - keep a backup if unsure
OUTPUT_CSV = "CP-scorebook.csv"     # also refreshed so score.py stays in sync

UNCHECKED_MARKERS = ("not yet checked", "not yet researched", "not yet confirmed", "")

# Column layout (1-indexed, matches the fixed sheet structure)
COL = {
    "brand": 2,
    "complaints_raw": 3,

    "hiring_raw": 4,
    "leadership_raw": 5,
    "growth_raw": 6,
    "vendor_raw": 7,
    "complaint_tier": 10,
    "leadership_tier": 11,
    "techstack_tier": 12,
    "hiring_tier": 13,
    "growth_tier": 14,
    "icp_fit": 15,
}

# Same rubric text used by tier_classifier.py - the RULES are Mimi's, already
# decided earlier in this project. The LLM only applies them, doesn't invent them.
COMPLAINTS_RUBRIC = """Classify the complaint severity using EXACTLY these rules, nothing else:
- S (strong): complaints specifically about returns/refunds/post-purchase, corroborated
  across 2+ independent sources (e.g. Trustpilot AND BBB), recent (last 6-12 months)
- M (moderate): same kind of complaint, but from only ONE credible source

- W (weak): only 1 isolated complaint, OR complaints are old/resolved, OR complaints
  are about product quality rather than returns/refunds specifically
- N (none): no returns/refund-specific complaints found
Reply with ONLY one letter: S, M, W, or N. No explanation, no punctuation."""

TECHSTACK_RUBRIC = """Classify the current returns vendor status using EXACTLY these rules:
- C: they use a competitor returns platform (Loop, AfterShip, Redo, Onward, Happy Returns)
  AND there are complaints about the returns process specifically
- H: they use a competitor returns platform AND customers seem satisfied, OR no
  returns-specific complaints were found
- NT: no third-party returns platform found - manual or in-house process
Reply with ONLY: C, H, or NT. No explanation, no punctuation."""

LEADERSHIP_RUBRIC = """Classify using EXACTLY this rule:
- Y: a CX Head, COO, CTO, or Founder-level hire or return happened within roughly
  the last 12 months
- N: no such hire found, OR it happened more than ~12 months ago
Reply with ONLY: Y or N. No explanation, no punctuation."""

HIRING_RUBRIC = """Classify open job postings using EXACTLY these rules:
- CL (clustered): 3 or more relevant open roles (Returns Manager, Logistics Ops, CX Platform)
- P (present): 1-2 relevant open roles
- N (none): no relevant open roles found
Reply with ONLY: CL, P, or N. No explanation, no punctuation."""

GROWTH_RUBRIC = """Classify using EXACTLY this rule:
- Y: the text describes real growth (funding raised, new stores/retail expansion,
  geographic/international expansion, new product lines)
- N: no growth signal found, OR the text describes decline/financial distress instead
  of growth
Reply with ONLY: Y or N. No explanation, no punctuation."""


def is_red_fill(cell):
    """
    Returns True if the cell's background fill looks red/pink. Excel stores
    fill color as an 8-char ARGB hex (e.g. 'FFFDDFE3' for a pale pink
    highlight, or 'FFFF0000' for pure red). Checks fgColor (not start_color,
    which can be unreliable depending on how the fill was applied) and uses
    a loose enough threshold to catch pale/pastel red highlights, not just
    fully saturated red.
    """
    fill = cell.fill
    if not fill or fill.patternType != "solid":
        return False
    color = fill.fgColor
    if not color or not isinstance(color.rgb, str) or len(color.rgb) != 8:
        return False
    try:
        r = int(color.rgb[2:4], 16)
        g = int(color.rgb[4:6], 16)
        b = int(color.rgb[6:8], 16)
    except ValueError:
        return False
    # "Red-ish": red is the dominant channel, noticeably higher than both
    # green and blue - catches pale pink highlights as well as bright red.
    return r > 200 and (r - g) > 15 and (r - b) > 15


def call_ollama(system_prompt, user_prompt):
    if not USE_LLM_EXPLANATIONS:
        return None
    payload = {
        "model": OLLAMA_MODEL,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        "stream": False,
    }
    req = urllib.request.Request(
        OLLAMA_URL,
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            data = json.loads(resp.read().decode("utf-8"))
            return data["message"]["content"].strip()
    except Exception as e:
        print(f"  [Ollama call failed - {type(e).__name__}: {e}]")
        return None


def classify(rubric_prompt, raw_text, valid_codes):
    """Calls the LLM and validates the response is an allowed code. Returns None on anything else."""
    result = call_ollama(rubric_prompt, f"Raw research text: \"{raw_text}\"")
    if not result:
        return None
    cleaned = result.strip().upper().rstrip(".")
    if cleaned in valid_codes:
        return cleaned
    for code in valid_codes:
        if cleaned == code or cleaned.startswith(code + " ") or cleaned.startswith(code + "."):
            return code
    print(f"  [Unparseable response '{result}' - leaving tier blank rather than guessing]")
    return None


def is_unchecked(text):
    return (text or "").strip().lower() in UNCHECKED_MARKERS


def main(overwrite=False):
    if USE_LLM_EXPLANATIONS:
        print(f"Using local Ollama ({OLLAMA_MODEL}) for classification.\n")

    if not USE_LLM:
        print("USE_LLM_EXPLANATIONS is not set to true - skipping AI classification. "
              "Set it and run 'ollama serve' to enable it.")
        print("Still exporting the current spreadsheet to CSV so score.py can run...")
        wb = openpyxl.load_workbook(INPUT_XLSX)
        ws = wb[wb.sheetnames[0]]
        with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
            import csv
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)
        print(f"Exported {OUTPUT_CSV} (tiers unchanged from the spreadsheet).")
        return

    wb = openpyxl.load_workbook(INPUT_XLSX)
    ws = wb[wb.sheetnames[0]]

    changed, skipped_red = 0, 0

    for r in range(2, ws.max_row + 1):
        brand = ws.cell(row=r, column=COL["brand"]).value
        if not brand:
            continue

        icp_fit = str(ws.cell(row=r, column=COL["icp_fit"]).value or "").strip().upper()
        if icp_fit == "N":
            print(f"{brand}: skipped - Product_Fits_Returns_Category = N (not a real ICP fit, e.g. food/beverage)")
            continue

        jobs = [
            ("complaints_raw", "complaint_tier", COMPLAINTS_RUBRIC, {"S", "M", "W", "N"}),
            ("vendor_raw", "techstack_tier", TECHSTACK_RUBRIC, {"C", "H", "NT"}),
            ("leadership_raw", "leadership_tier", LEADERSHIP_RUBRIC, {"Y", "N"}),
            ("hiring_raw", "hiring_tier", HIRING_RUBRIC, {"CL", "P", "N"}),
            ("growth_raw", "growth_tier", GROWTH_RUBRIC, {"Y", "N"}),
        ]

        for raw_col_key, tier_col_key, rubric, valid_codes in jobs:
            raw_cell = ws.cell(row=r, column=COL[raw_col_key])
            tier_cell = ws.cell(row=r, column=COL[tier_col_key])

            if is_red_fill(raw_cell) or is_red_fill(tier_cell):
                skipped_red += 1
                continue

            current_tier = str(tier_cell.value or "").strip()
            raw_text = raw_cell.value

            if (overwrite or not current_tier) and not is_unchecked(raw_text):
                result = classify(rubric, raw_text, valid_codes)
                if result:
                    tier_cell.value = result
                    changed += 1
                    print(f"{brand}: {tier_col_key} -> {result}")

    wb.save(OUTPUT_XLSX)

    # Refresh the CSV too, so score.py stays in sync with this run
    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
        import csv
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)

    print(f"\nDone. {changed} tier cells classified. {skipped_red} cells skipped (red-highlighted).")
    print(f"Saved to {OUTPUT_XLSX} and refreshed {OUTPUT_CSV}.")


if __name__ == "__main__":
    overwrite = "--overwrite" in sys.argv
    main(overwrite=overwrite)